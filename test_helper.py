# _*_ coding:utf-8 _*_

import random
import re
from copy import copy
from openpyxl import load_workbook

data_res = r'C:\Users\Administrator\.jenkins\workspace\psimsendemail\untitled\report\test_report.xlsx'
data = r'C:\Users\Administrator\.jenkins\workspace\psimsendemail\untitled\data\register.xlsx'

def get_phonenum():
    '''自动生成手机号码'''
    phone = '1'+random.choice(['3','4','5','7','8','9'])
    for i in range(9):
        num = random.randint(0,9)
        phone += str(num)
    return phone

def copy_sheet(sheetname):
    '''
    跨文件复制sheet
    :param sheetname:
    :return:
    '''
    # 追加内容来自于该文件
    tpl_sheet_name = sheetname
    wb_tpl = load_workbook(data)
    ws_tpl = wb_tpl[tpl_sheet_name]

    # 要保留的文件
    wb = load_workbook(data_res)
    # active会选择你打开时选中的sheet表，如果有多个sheet表不确定时，建议使用wb['sheet_name']
    # ws = wb.active
    ws = wb.create_sheet(f'{sheetname}')
    max_row = ws.max_row-1

    cur_row = 0
    for row in ws_tpl:
      cur_row += 1
      for cell in row:
        # 复制值
        # ws[cell.coordinate].value = cell.value
        try:
          pos = f'{cell.column_letter}{max_row + cur_row}'
        except Exception as e:
          continue
        ws[pos].value = cell.value
        ws[pos].data_type = copy(cell.data_type)

        if cell.has_style:
          # 复制字体，填充，对齐，边框
          ws[pos].font = copy(cell.font)
          ws[pos].fill = copy(cell.fill)
          ws[pos].number_format = copy(cell.number_format)
          ws[pos].alignment = copy(cell.alignment)
          ws[pos].border = copy(cell.border)

          if cell.comment:
            # 复制备注
            ws[pos].comment = copy(cell.comment)

    # 复制合并单元格
    for x in ws_tpl.merged_cells.ranges:
      match = re.search(r'([a-zA-Z]+)(\d+):([a-zA-Z]+)(\d+)', x.coord)
      col_1, row_1, col_2, row_2 = match.groups()
      new_coord = f'{col_1}{int(row_1) + max_row}:{col_2}{int(row_2) + max_row}'
      ws.merge_cells(range_string=new_coord)

    # 保存
    wb.save(data_res)

def copy_sheet1():
    '''
    同一文件内复制sheet
    :return:
    '''
    workbook = load_workbook(data)
    sheet = workbook['Sheet1']
    target = workbook.copy_worksheet(sheet)
    workbook.save(data)
